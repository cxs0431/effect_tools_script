# coding=utf-8

from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook(filename='c:\\Users\\chenxingshun\\Downloads\\20180519\\SQL\\统一账户累积升级文档整理.xlsx', read_only=True)
ws = wb['Sheet3']
i=1
deslist = []
convertlist = [{}]
currindex = 0
desNo = ''
for row in ws.rows:
    ds = ws.cell(row=i, column = 2)
    i=i+1
    if str(ds.value).find('XQ')>-1:
        if ds.value != desNo and desNo != '':
            deslist += [(desNo, currindex,i)]
        currindex = i
        desNo = ds.value
i=1
tmp = ''
print(deslist)


dest_filename = 'empty_book.xlsx'
wb1 = Workbook()#filename = dest_filename, read_only = False)
ws1 = wb1.create_sheet(title="Pi")

ds_xq= ''
ds_fx =''
ds_sx =''
ds_yx=''
ds_cs=''

r = 0
for (des,start,end) in deslist:
    r = r + 1
    index = start
    while index < end:
        ds = ws.cell(row=index,column= 3)
        if (str(ds.value).find('需求描述') > -1):
            tmp = ''
        elif (str(ds.value).find('需求分析') > -1):
            ds_xq = tmp
            tmp = ''
        elif (str(ds.value).find('实现概要') > -1):   
            ds_fx = tmp
            tmp = ''
        elif (str(ds.value).find('影响范围') > -1):   
            ds_sx = tmp
            tmp = ''
        elif (str(ds.value).find('测试要点') > -1):   
            ds_yx = tmp
            tmp = ''
        else:
            tmp += str(ds.value)
        
        index = index + 1
    ds_cs=tmp
    _ = ws1.cell(column=1, row=r, value=des)
    _ = ws1.cell(column=2, row=r, value=ds_xq)
    _ = ws1.cell(column=3, row=r, value=ds_fx)
    _ = ws1.cell(column=4, row=r, value=ds_sx)
    _ = ws1.cell(column=5, row=r, value=ds_yx)
    _ = ws1.cell(column=6, row=r, value=ds_cs)
    
    

wb1.save(filename = dest_filename)
