#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv,configparser, os, time,datetime

from openpyxl import load_workbook
from openpyxl import Workbook

max_sys_monitor_col = 20
def init_time_col():
        timecol =['TIME',] ;
        d = datetime.datetime(2018, 9, 9,8,30,0)
        for x in range(max_sys_monitor_col - 1):
                delta = datetime.timedelta(seconds=15)
                d += delta
                timecol.append(d.strftime('%H:%M:%S'))
        return timecol;

def computer_sys_per(ws):
        row_title = ['名称','主机类型','CPU平均','CPU最高','时间点']
        ws.append(row_title);
        row = [];
        row.append('6','统一账户','虚拟机','','','')
        row.append('6','统一账户','虚拟机','','','')
        row.append('0','总控','实体机','=AVERAGE(各系统采集数据!C3:C1686)/100','')
        ws.append(row)
        
def parse_sys_monitor(ws, items ):
        if ( 'path' not in items):
                return;
        
        filelist = [];
        content=list();
        content.append(init_time_col());
        filelist = getDirCsv(items['path']);
        for p,f in filelist :
                content.append(readcsv(p+f))
        r = 0
        c = 0;
        print(len(content),len(content[0]))
        for col_index, col in enumerate(content):
                for row_index, cell in enumerate(col):
                        _ = ws.cell(column=col_index+1, row=row_index+1, value=cell)
                       
               
                                     
        

def getDirCsv( p ):
	p = os.getcwd() +str( p)
	if p=="":
		return [ ]
	p = p.replace( "/","\\")
	if p[ -1] != "\\":
		p = p+"\\"
	a = os.listdir( p )
	b = []
	for x in a:
		if os.path.isdir( p + x ):
			b += getDirList( p + x )
	b += [ (p , x)   for x in a if x.find('.csv') >- 1 ]
	return b


def readcsv( xlsfile, colname = 'Processor Time' ):
        listCol = []
        colindex = -1
        f = open(xlsfile)
        csv.register_dialect('colons', delimiter=',')
        reader = csv.reader(f, dialect='colons')
        for row_index, row in enumerate(reader):                
                if (row_index==1):
                        continue;
                
                if (row_index > max_sys_monitor_col) :
                        break;

                for column_index, cell in enumerate(row):
                        if str(cell).find(colname)>-1 and colindex == -1:
                                #print(row_index,column_index,cell)
                                listCol.append(cell)
                                colindex = column_index
                                break;
                        if (colindex == column_index):
                                #print(row_index,column_index,cell)
                                listCol.append(cell)

        return listCol

    
def func_convert(funcname):
        if (funcname == 'parse_sys_monitor'):
                return parse_sys_monitor;
        elif(funcname == 'computer_sys_per'):
                return computer_sys_per;
        return parse_sys_monitor;



curr_date =  time.strftime("%Y%m%d", time.localtime()) 
dest_filename = '主机性能及容量'+ curr_date +'.xlsx'

config = configparser.ConfigParser()
config.read ('report_config.cfg')

wb = Workbook()
wb.remove(wb['Sheet'])
for sheet in config.sections():
        for key in config[sheet]:
                config[sheet][key] = config[sheet][key].replace( "DATE",curr_date)

        ws1 = wb.create_sheet(title=sheet)
        #print(config[sheet]['func'])
        func_convert(config[sheet]['func'])(ws1, config[sheet])
wb.save(filename = dest_filename)

#listCol = readcsv('.\\data\\20180907\\zrtperform.csv')
#print(listCol,sep='\n')
